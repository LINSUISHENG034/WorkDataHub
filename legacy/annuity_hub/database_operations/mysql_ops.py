# /database_operations/mysql_ops.py

import pandas as pd
import re
from datetime import datetime
from dataclasses import dataclass, field
from sqlalchemy import create_engine, inspect
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from tqdm import tqdm
from textwrap import dedent
from openpyxl import load_workbook
from contextlib import suppress

from config_manager.config_manager import ConfigManager
from logger.logger import logger as logger


@dataclass
class MySqlDBManager:
    # 数据库连接属性
    database: str = None  # 将数据库设置为可选参数
    config: ConfigManager = field(default_factory=ConfigManager)
    engine: any = field(init=False, default=None)
    conn: any = field(init=False, default=None)
    cursor: any = field(init=False, default=None)

    def __post_init__(self):
        self.connect()

    def connect(self):
        config = self.config.get_mysql_config()
        # 根据 database 是否为 None 构建不同的 engine_url
        if self.database:
            engine_url = (
                f"mysql+pymysql://{config['user']}:{config['pass']}@"
                f"{config['host']}:{config['port']}/{self.database}"
                f"?charset={config['charset']}"
            )
        else:
            engine_url = (
                f"mysql+pymysql://{config['user']}:{config['pass']}@"
                f"{config['host']}:{config['port']}"
                f"?charset={config['charset']}"
            )
        try:
            # 在 create_engine 中追加 pool_pre_ping 和 connect_args 参数
            self.engine = create_engine(
                engine_url,
                pool_size=config.get("pool_size", 10),
                max_overflow=20,
                pool_pre_ping=True,
                connect_args={"connect_timeout": 30},
            )
            self.conn = self.engine.raw_connection()
            self.cursor = self.conn.cursor()
        except SQLAlchemyError as e:
            logger.error(f"Error connecting to database: {e}")
            raise

    def close(self):
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
        if self.engine:
            self.engine.dispose()

    def __enter__(self):
        if not self.engine:
            self.connect()
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    @staticmethod
    def _validate_table_name(table_name: str):
        if not re.match(r"^[\w]+$", table_name):
            raise ValueError(f"Invalid table name: {table_name}")

    @staticmethod
    def _prepare_data(data: pd.DataFrame):
        # 使用 replace 将所有 NaN 替换为 None，这是 MySQL 中的 NULL
        return data.replace({pd.NA: None, pd.NaT: None, float("nan"): None})

    def _insert_data(
        self,
        table_name: str,
        data: pd.DataFrame,
        batch_size: int,
        ignore_conflicts=False,
    ):
        try:
            # 分批插入数据
            for start_row in range(0, data.shape[0], batch_size):
                end_row = min(start_row + batch_size, data.shape[0])
                batch_data = data.iloc[start_row:end_row]

                records = batch_data.to_dict(orient="records")
                columns = ", ".join(
                    [f"`{col}`" for col in batch_data.columns]
                )  # 字段名使用反引号包围
                placeholders = ", ".join(["%s"] * len(batch_data.columns))

                insert_clause = f"INSERT {'IGNORE' if ignore_conflicts else ''} INTO `{table_name}` ({columns}) VALUES ({placeholders})"

                # 批量执行插入
                self.cursor.executemany(
                    insert_clause,
                    [tuple(r[col] for col in batch_data.columns) for r in records],
                )

        except IntegrityError as e:
            logger.error(f"Integrity error during data insert: {e}")
            self.conn.rollback()
            raise
        except SQLAlchemyError as e:
            logger.error(f"SQLAlchemy error during data insert: {e}")
            self.conn.rollback()
            raise
        except Exception as e:
            logger.error(f"General error during data insert: {e}")
            self.conn.rollback()
            raise

    def import_data(
        self,
        table_name: str,
        data: pd.DataFrame,
        delete_existing: bool = False,
        batch_size: int = 1000,
    ):
        """
        直接插入数据，遇到主键冲突会抛出异常。

        参数:
        - table_name (str): 表名。
        - data (pd.DataFrame): 要插入的数据。
        - delete_existing (bool, optional): 是否删除现有的数据。默认是 False。
        - batch_size (int, optional): 每批插入的数据大小。默认是 1000。
        """
        self._validate_table_name(table_name)
        data = self._prepare_data(data)

        try:
            if delete_existing:
                self.cursor.execute(f"TRUNCATE TABLE {table_name}")
                self.conn.commit()

            self.conn.begin()  # 开启事务
            self._insert_data(
                table_name, data, batch_size, ignore_conflicts=False
            )  # 不忽略主键冲突
            self.conn.commit()  # 提交事务
            # logger.info(f"Data successfully imported to {table_name}.")
        except Exception as e:
            self.conn.rollback()  # 发生错误时回滚
            logger.error(f"Error in import_data: {e}")
            raise

    def import_data_with_ignore(
        self, table_name: str, data: pd.DataFrame, batch_size: int = 1000
    ):
        """
        插入数据，跳过主键冲突
        参数:
        - table_name (str): 表名。
        - data (pd.DataFrame): 要插入的数据。
        - batch_size (int, optional): 每批插入的数据大小。默认是1000。
        """
        self._validate_table_name(table_name)
        data = self._prepare_data(data)

        try:
            self.conn.begin()  # 开启事务
            self._insert_data(table_name, data, batch_size, ignore_conflicts=True)
            self.conn.commit()  # 提交事务
            logger.info(f"Data imported to {table_name} with conflicts ignored.")
        except Exception as e:
            self.conn.rollback()
            logger.error(f"Error in import_data_with_ignore: {e}")
            raise

    def import_data_with_update_on_conflict(
        self,
        table_name: str,
        data: pd.DataFrame,
        update_field: str,
        batch_size: int = 1000,
    ):
        """
        插入数据，发生主键冲突时，根据指定的字段（如updateTime）决定是否更新数据。

        参数:
        - table_name (str): 表名。
        - data (pd.DataFrame): 要插入的数据。
        - update_field (str): 在主键冲突时用来决定是否更新数据的字段，通常是时间字段。
        - batch_size (int, optional): 每批插入的数据大小。默认是1000。
        """
        self._validate_table_name(table_name)
        data = self._prepare_data(data)

        try:
            self.conn.begin()  # 开启事务
            for start_row in range(0, data.shape[0], batch_size):
                end_row = min(start_row + batch_size, data.shape[0])
                batch_data = data.iloc[start_row:end_row]

                records = batch_data.to_dict(orient="records")
                columns = ", ".join([f"`{col}`" for col in batch_data.columns])
                placeholders = ", ".join(["%s"] * len(batch_data.columns))

                # 插入语句，同时在主键冲突时更新指定字段
                insert_clause = f"""
                    INSERT INTO `{table_name}` ({columns})
                    VALUES ({placeholders})
                    ON DUPLICATE KEY UPDATE
                    `{update_field}` = IF(VALUES(`{update_field}`) > `{update_field}`, VALUES(`{update_field}`), `{update_field}`)
                """

                # 批量执行插入和更新
                self.cursor.executemany(
                    insert_clause,
                    [tuple(r[col] for col in batch_data.columns) for r in records],
                )

            self.conn.commit()  # 提交事务
            logger.info(
                f"Data imported to {table_name} with update on conflict based on {update_field}."
            )
        except Exception as e:
            self.conn.rollback()  # 发生错误时回滚
            logger.error(f"Error in import_data_with_update_on_conflict: {e}")
            raise

    def add_foreign_key(self, table_name: str):
        # 使用 SQLAlchemy 的 inspector 来检查当前数据库
        inspector = inspect(self.engine)
        # 获取当前表已经存在的外键
        current_keys = self.get_foreign_keys(table_name)
        # 定义外键映射关系
        foreign_keys_mapping = {
            "机构代码": {"fk_database": "mapping", "fk_table": "组织架构"},
            "产品线代码": {"fk_database": "mapping", "fk_table": "产品线"},
            "年金计划号": {
                "fk_database": "mapping",
                "fk_table": "年金计划",
                "alias": ["计划代码", "计划号"],
            },
            "组合代码": {
                "fk_database": "mapping",
                "fk_table": "组合计划",
                "alias": ["组合计划代码", "组合编号"],
            },
            "产品ID": {"fk_database": "mapping", "fk_table": "产品明细"},
            "指标编码": {"fk_database": "mapping", "fk_table": "利润指标"},
        }

        # 提取表中所有列的名称
        columns = [col["name"] for col in inspector.get_columns(table_name)]

        for index, mapping in foreign_keys_mapping.items():
            # 如果外键表与当前表相同，则跳过
            if table_name == mapping["fk_table"]:
                continue

            # 将别名和主键名组合起来，检查这些名称是否存在于当前表的列中
            potential_keys = mapping.get("alias", []) + [index]
            # 使用生成器表达式找到第一个匹配的列名
            column_name = next(
                (name for name in potential_keys if name in columns), None
            )

            if column_name:
                # 构造外键约束的名称
                constraint_name = f"FK_{mapping['fk_table']}_{table_name}"

                # 如果外键约束已经存在，则记录日志并跳过
                if constraint_name in current_keys:
                    logger.info(
                        f"Constraint '{constraint_name}' for column '{column_name}' already exists. Skipping."
                    )
                    continue

                # 构造添加外键的 SQL 语句
                fk_sql = f"""
                    ALTER TABLE `{table_name}` ADD CONSTRAINT `{constraint_name}` FOREIGN KEY (`{column_name}`) 
                    REFERENCES `{mapping["fk_database"]}`.`{mapping["fk_table"]}`(`{index}`) 
                    ON DELETE RESTRICT ON UPDATE CASCADE;
                """
                try:
                    # 执行 SQL 语句，添加外键
                    self.cursor.execute(fk_sql)
                    self.conn.commit()
                    logger.info(
                        f"Added foreign key constraint '{constraint_name}' to '{table_name}' for column '{column_name}'."
                    )
                except SQLAlchemyError as e:
                    # 如果执行过程中出错，则回滚并记录错误
                    self.conn.rollback()
                    logger.error(
                        f"Error adding foreign key to '{table_name}' for column '{column_name}': {e}"
                    )

    def add_index_key(self, index_keys: list, exclude_tables: list = None):
        if exclude_tables is None:
            exclude_tables = []

        self.cursor.execute("SHOW TABLES")
        tables = self.cursor.fetchall()
        for table in tqdm(tables, desc="Adding index key"):
            table_name = table[0]
            if table_name not in exclude_tables:
                for index_key in index_keys:
                    if self.column_exists(table_name, index_key):
                        index_name = f"IDX_{table_name}_{index_key}"
                        if not self.index_exists(table_name, index_name):
                            try:
                                alter_sql = f"CREATE INDEX `{index_name}` ON `{table_name}` (`{index_key}`);"
                                self.cursor.execute(alter_sql)
                            except Exception as e:
                                self.conn.rollback()
                                raise RuntimeError(
                                    f"Error altering table {table_name}: {e}"
                                )
                        else:
                            logger.info(
                                f"Index {index_name} already exists on {table_name}."
                            )
                    else:
                        logger.warning(
                            f"Column {index_key} does not exist in table {table_name}."
                        )

    def column_exists(self, table_name: str, column_name: str) -> bool:
        try:
            query = """
            SELECT COUNT(1) 
            FROM information_schema.columns 
            WHERE table_schema = %s 
              AND table_name = %s 
              AND column_name = %s;
            """
            self.cursor.execute(query, (self.database, table_name, column_name))
            return self.cursor.fetchone()[0] > 0
        except Exception as e:
            logger.error(f"Error checking column existence: {e}")
            raise

    def delete_rows_by_criteria(
        self,
        table_name: str,
        criteria: pd.DataFrame,
        fields: list,
        batch_size: int = 100,
    ):
        """
        根据多个字段的唯一组合从MySQL表中删除行。

        Args:
        - table_name (str): 要删除行的表名。
        - criteria (pd.DataFrame): 包含唯一组合的DataFrame。
        - fields (list): 需要匹配的字段列表。
        - batch_size (int, optional): 每批次删除的大小。默认是100。

        Returns:
        int: 被删除的行数。
        """
        total_deleted = 0
        if criteria.empty:
            return total_deleted

        for i in range(0, len(criteria), batch_size):
            batch_criteria = criteria.iloc[i : i + batch_size]

            for _, row in batch_criteria.iterrows():
                conditions = []
                values = []

                for field in fields:
                    if isinstance(row[field], (pd.Timestamp, datetime)):
                        # 将日期字段转换为 "YYYY-MM" 格式并进行条件匹配
                        conditions.append(f"DATE_FORMAT({field}, '%%Y-%%m') = %s")
                        values.append(row[field].strftime("%Y-%m"))
                    else:
                        conditions.append(f"{field} = %s")
                        values.append(row[field])

                delete_statement = f"DELETE FROM {table_name} WHERE " + " AND ".join(
                    conditions
                )

                try:
                    # 这里将条件的参数用元组形式传递给execute
                    self.cursor.execute(delete_statement, tuple(values))
                    total_deleted += self.cursor.rowcount
                    self.conn.commit()
                except SQLAlchemyError as e:
                    self.conn.rollback()
                    raise

        return total_deleted

    def delete_rows_by_months(self, table_name: str, months: list, date_col="日期"):
        """
        Delete rows from a MySQL table based on a list of months for a given date column.

        Args:
        - table_name (str): The name of the table from which rows should be deleted.
        - months (list): A list of months in the format "yyyy-mm" to identify rows to be deleted.
        - date_col (str, optional): The name of the date column to use for filtering. Default is '日期'.

        Returns:
        int: The total number of rows deleted.

        Note:
        Ensure that the months in the provided list are in the "yyyy-mm" format.
        """
        total_deleted = 0  # Counter for the total number of rows deleted
        if not months:
            return
        for month in months:
            if re.match(r"\d{4}-\d{2}", month):
                try:
                    delete_statement = f'DELETE FROM {table_name} WHERE DATE_FORMAT({date_col}, "%%Y-%%m") = %s'
                    self.cursor.execute(delete_statement, (month,))
                    total_deleted += (
                        self.cursor.rowcount
                    )  # Update the counter with the number of rows deleted
                except SQLAlchemyError as e:
                    self.conn.rollback()  # Rollback the transaction if an error occurs
                    raise  # Re-raise the exception to be handled by the caller
        self.conn.commit()  # Commit the changes only if all delete operations were successful
        return total_deleted  # Return the total number of rows deleted

    def delete_rows_by_values(
        self, table_name: str, values: list, col_name="保单号", batch_size=100
    ):
        """
        Delete rows from a database table based on unique values from a DataFrame column in batches.

        Args:
        - table_name (str): The name of the table from which rows should be deleted.
        - values (list): The list of values to be deleted.
        - col_name (str): The column name to match for deletion.
        - batches (int): Number of values to delete in each batch.

        Returns:
        int: The total number of rows deleted.
        """
        total_deleted = 0  # Counter for the total number of rows deleted

        for i in range(0, len(values), batch_size):
            batch_values = values[i : i + batch_size]
            placeholders = ",".join(["%s"] * len(batch_values))
            delete_statement = (
                f"DELETE FROM {table_name} WHERE {col_name} IN ({placeholders})"
            )

            try:
                self.cursor.execute(delete_statement, batch_values)
                total_deleted += self.cursor.rowcount
                self.conn.commit()
            except SQLAlchemyError as e:
                self.conn.rollback()
                raise

        return total_deleted  # Return the total number of rows deleted

    def get_column_names(self, table_name: str):
        self.cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = [column[0] for column in self.cursor.fetchall()]
        return columns

    def get_columns_info_from_all_tables(self):
        inspector = inspect(self.engine)
        table_names = inspector.get_table_names()
        dfs = []

        for table_name in tqdm(table_names, desc="Fetching Columns"):
            infos = inspector.get_columns(table_name)
            df = pd.DataFrame(infos)
            df.rename(columns={"name": "column_name"}, inplace=True)
            df["table_name"] = table_name
            df["database_name"] = (
                self.database
            )  # Use self.database_name instead of hard-coded value
            dfs.append(df)

        return pd.concat(dfs, ignore_index=True)

    def get_foreign_keys(self, table_name: str):
        inspector = inspect(self.engine)
        foreign_keys = inspector.get_foreign_keys(table_name)
        return [fk["name"] for fk in foreign_keys if "name" in fk]

    def get_full_table(self, table_name: str = None, sql_statement: str = None):
        try:
            if table_name:
                sql_statement = f"SELECT * FROM {table_name}"

            if not sql_statement:
                raise ValueError("Either table_name or sql_statement must be provided")

            # 执行查询
            self.cursor.execute(sql_statement)
            data = self.cursor.fetchall()

            # 获取列名
            columns = [desc[0] for desc in self.cursor.description]

            # 检查数据是否为空
            if not data:
                return pd.DataFrame(columns=columns)

            df = pd.DataFrame(data, columns=columns)
            return df

        except SQLAlchemyError as e:
            logger.error(f"Error fetching data: {e}")
            return pd.DataFrame()  # 返回空的DataFrame以防止程序崩溃
        except ValueError as ve:
            logger.error(ve)
            return pd.DataFrame()  # 返回空的DataFrame以防止程序崩溃

    def index_exists(self, table_name: str, index_name: str) -> bool:
        try:
            query = """
            SELECT COUNT(1) 
            FROM information_schema.statistics 
            WHERE table_schema = %s 
              AND table_name = %s 
              AND index_name = %s;
            """
            self.cursor.execute(query, (self.database, table_name, index_name))
            return self.cursor.fetchone()[0] > 0
        except Exception as e:
            logger.error(f"Error checking index existence: {e}")
            raise

    def show_table_rows(self):
        """
        Fetch the number of rows and columns for each table in the database.

        Returns:
        - table_info: A dictionary where the key is the table name and the
          value is a list containing two numbers:
          [number_of_rows, number_of_columns]
        """

        table_infos = {}

        try:
            self.cursor.execute("SHOW TABLES")
            tables = self.cursor.fetchall()

            for table in tables:
                table_name = table[0]
                # Fetching the row count
                self.cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = self.cursor.fetchone()[0]

                # Fetching the column count
                query = """
                    SELECT COUNT(*)
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """
                self.cursor.execute(query, (self.database, table_name))
                col_count = self.cursor.fetchone()[0]

                table_infos[table_name] = [count, col_count]
        except Exception as e:
            self.conn.rollback()
            raise RuntimeError(f"Error fetching table process: {e}")

        return table_infos

    def switch_database(self, new_database):
        self.database = new_database
        if self.engine:
            self.engine.dispose()  # 释放旧的数据库引擎
        self.connect()  # 重新连接到新的数据库

    def table_exists(self, table_name: str):
        query = "SELECT COUNT(*) FROM information_schema.tables WHERE table_name = %s"
        self.cursor.execute(query, (table_name,))
        result = self.cursor.fetchone()[0]
        return result == 1

    def uniform_all_tables(self):
        statement = f"""
            SELECT TABLE_NAME FROM information_schema.tables WHERE TABLE_SCHEMA='{self.database}';
        """
        try:
            self.cursor.execute(statement)
            rows = self.cursor.fetchall()
            for row in rows:
                table_name = row[0]
                sql_statement = dedent(f"""
                    ALTER TABLE {table_name}
                    ADD COLUMN id INT NOT NULL AUTO_INCREMENT FIRST,
                    ADD CONSTRAINT row_index UNIQUE (id);""")
                self.cursor.execute(sql_statement)
            self.conn.commit()
        except Exception as e:
            # Rollback the transaction in case of error
            self.conn.rollback()
            # Raise the exception as a RuntimeError
            logger.info(f"Error processing tables: {e}")

    def convert_float_to_double(self):
        try:
            # 查找所有float类型的列
            query = """
            SELECT TABLE_NAME, COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE DATA_TYPE = 'float' AND TABLE_SCHEMA = %s
            """
            self.cursor.execute(query, (self.database,))
            columns_to_alter = self.cursor.fetchall()

            # 遍历结果并更改列类型
            for table_name, column_name in columns_to_alter:
                alter_statement = (
                    f"ALTER TABLE {table_name} MODIFY {column_name} DOUBLE"
                )
                self.cursor.execute(alter_statement)
                logger.info(f"Modified {column_name} in {table_name} to DOUBLE.")

            self.conn.commit()
            logger.info("All float columns have been converted to double successfully.")

        except Exception as e:
            self.conn.rollback()
            logger.error(f"Error during converting float columns to double: {e}")
            raise

    def create_table(self, table_name, data: pd.DataFrame, primary_key: str = None):
        # 检查表格是否存在
        if self.table_exists(table_name):
            logger.error(f"Table `{table_name}` already exists.")
            return

        # 开始构建SQL语句
        sql = f"CREATE TABLE `{table_name}` (\n"

        # 遍历DataFrame的列，为每个列生成字段定义
        df_head_100 = data.head(100)
        for column, dtype in df_head_100.dtypes.items():
            sql_type = self.dtype_to_sql(dtype)
            sql += f"  `{column}` {sql_type},\n"

        # 移除最后一个逗号并添加结束的括号
        sql = sql.rstrip(",\n") + "\n);"

        # 执行SQL创建表格语句
        try:
            self.cursor.execute(sql)
            self.conn.commit()
            logger.info(f"Table `{table_name}` created successfully.")
        except SQLAlchemyError as e:
            self.conn.rollback()
            logger.error(f"Error creating table `{table_name}`: {e}")

        # 执行SQL创建表格主键
        if primary_key:
            try:
                alter_sql = (
                    f"ALTER TABLE `{table_name}` ADD PRIMARY KEY (`{primary_key}`);"
                )
                self.cursor.execute(alter_sql)
            except Exception as e:
                self.conn.rollback()
                raise RuntimeError(
                    f"Error altering table {table_name} to add primary key: {e}"
                )

    def summarize_customer_fields(
        self, target_table: str, target_database: str = "temp"
    ):
        """
        汇总指定数据库中所有表格中与客户名称相关的字段并插入到指定表格中。

        Args:
        - target_table (str): 插入的目标表名。
        - target_database (str): 插入的目标数据库名。
        """
        inspector = inspect(self.engine)
        table_names = inspector.get_table_names()
        cname_dict = {
            "企康缴费_健": "投保人",
            "企康缴费_养": "客户名称",
            "团养缴费": "企业名称",
            "手工调整": "客户名称",
            "提费扩面": "客户名称",
            "收入明细": "客户名称",
            "规模明细": "客户名称",
            "规模明细_回溯": "客户名称",
            "账管数据": "企业名称",
            "企年受托中标": "客户全称",
            "企年受托已客": "客户名称",
            "企年受托战客": "客户名称",
            "企年受托流失": "客户全称",
            "企年投资中标": "客户全称",
            "企年投资已客": "客户名称",
            "企年投资战客": "客户名称",
            "企年投资流失": "客户全称",
            "团养中标": "客户名称",
            "投资客户分摊比例表": "企业名称",
            "续签客户清单": "客户名称",
            "职年受托已客": "客户名称",
            "职年投资已客": "客户名称",
        }

        for table_name in tqdm(table_names, desc="Processing Tables"):
            if table_name in cname_dict:
                field = cname_dict[table_name]
                # 构造插入语句
                sql = dedent(f"""
                    INSERT INTO {target_database}.`{target_table}`(`客户名称`, `来源`)
                    (
                        SELECT t.`{field}` AS `客户名称`, '{table_name}' AS `数据来源`
                        FROM {self.database}.`{table_name}` AS t
                        WHERE t.company_id IS NULL
                        GROUP BY t.`{field}`
                    );
                """)
                try:
                    self.cursor.execute(sql)
                    self.conn.commit()
                    logger.info(f"Inserted data from {table_name} using field {field}.")
                except SQLAlchemyError as e:
                    self.conn.rollback()
                    logger.error(
                        f"Error inserting data from {table_name} using field {field}: {e}"
                    )

    @staticmethod
    def dtype_to_sql(dtype):
        """将pandas数据类型转换为SQL数据类型"""
        if pd.api.types.is_integer_dtype(dtype):
            return "INT"
        elif pd.api.types.is_float_dtype(dtype):
            return "FLOAT"
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            return "DATETIME"
        elif pd.api.types.is_bool_dtype(dtype):
            return "BOOLEAN"
        else:
            # 默认使用VARCHAR
            return "VARCHAR(255)"

    @staticmethod
    def read_data_from_excel(file_path, sheet_name="sheet0"):
        df = None
        if file_path.endswith(".csv"):
            try:
                local_data = pd.read_csv(
                    file_path, delimiter=",", encoding="utf-8-sig", dtype="str"
                )
                local_data = local_data.rename(columns=str.strip)
                df = local_data
            except UnicodeDecodeError:
                try:
                    local_data = pd.read_csv(
                        file_path, delimiter=",", encoding="cp936", dtype="str"
                    )
                    local_data = local_data.rename(columns=str.strip)
                    df = local_data
                except UnicodeDecodeError:
                    # Handle or log the error as needed
                    pass
        elif file_path.endswith((".xlsx", ".xls")):
            local_data = pd.read_excel(
                file_path, engine="openpyxl", sheet_name=sheet_name
            )
            df = local_data

        return df

    def add_company_id_column(self):
        cname_dict = {
            "企康缴费_健": "投保人",
            "企康缴费_养": "客户名称",
            "团养缴费": "企业名称",
            "手工调整": "客户名称",
            "提费扩面": "客户名称",
            "收入明细": "客户名称",
            "规模明细": "客户名称",
            "规模明细_回溯": "客户名称",
            "账管数据": "企业名称",
            "企年受托中标": "客户全称",
            "企年受托已客": "客户名称",
            "企年受托战客": "客户名称",
            "企年受托流失": "客户全称",
            "企年投资中标": "客户全称",
            "企年投资已客": "客户名称",
            "企年投资战客": "客户名称",
            "企年投资流失": "客户全称",
            "团养中标": "客户名称",
            "投资客户分摊比例表": "企业名称",
            "续签客户清单": "客户名称",
            "职年受托已客": "客户名称",
            "职年投资已客": "客户名称",
        }
        self.cursor.execute("SHOW TABLES")
        tables = [table[0] for table in self.cursor.fetchall()]
        for table_name in tqdm(tables, desc="Adding company_id column"):
            if table_name in cname_dict:
                customer_col = cname_dict[table_name]
                if not self.column_exists(table_name, "company_id"):
                    try:
                        # 添加 company_id 列和索引
                        self.cursor.execute(
                            f"ALTER TABLE `{table_name}` ADD COLUMN `company_id` VARCHAR(50) NULL;"
                        )
                        self.cursor.execute(
                            f"CREATE INDEX `KY_客户号` ON `{table_name}` (`company_id`);"
                        )
                        self.conn.commit()
                    except Exception as e:
                        self.conn.rollback()
                        logger.error(
                            f"在表 {table_name} 中添加 company_id 列时出错: {e}"
                        )
                        continue  # 跳过当前表，继续下一个表

                try:
                    # 更新客户名称中的括号
                    update_sql = f"""
                        UPDATE `{table_name}` AS t1
                        SET t1.`{customer_col}` = REPLACE(REPLACE(t1.`{customer_col}`, '(', '（'), ')', '）')
                        WHERE t1.`{customer_col}` LIKE '%(%' OR t1.`{customer_col}` LIKE '%)%';
                    """
                    self.cursor.execute(update_sql)

                    # 通过关联 annuity_cname_mapping 表更新 company_id
                    alter_sql = f"""
                        UPDATE `{table_name}` AS t1
                        INNER JOIN `enterprise`.`annuity_cname_mapping` AS t2
                        ON t1.`{customer_col}` = t2.`客户名称`
                        SET t1.`company_id` = t2.`company_id`
                        WHERE t1.`company_id` IS NULL;
                    """
                    self.cursor.execute(alter_sql)
                    self.conn.commit()
                except Exception as e:
                    self.conn.rollback()
                    logger.error(f"在表 {table_name} 中更新 company_id 时出错: {e}")


if __name__ == "__main__":
    selected_sub_process = int(input("Please enter an integer: "))

    # 删除数据
    if selected_sub_process == 1:
        with MySqlDBManager(database="") as mysqldb:
            mysqldb.uniform_all_tables()

    # 读取Excel表格并写入数据库
    elif selected_sub_process == 2:
        with MySqlDBManager(database="other") as mysqldb:
            excel_file_path = r"D:\Users\linsuisheng034\Desktop\IDs.xlsx"
            for table_name in ["data"]:
                sheet_name = table_name
                mysql_table_name = "查询名单"
                # 读取数据
                data = mysqldb.read_data_from_excel(excel_file_path, sheet_name)

                # 数据清洗1
                # from common_utils.common_utils import clean_company_name
                # data['客户名称'] = data['原客户名称'].apply(clean_company_name)

                # 数据清洗2
                # from data_handler.data_cleaner import AnnuityPerformanceCleaner
                # cleaner = AnnuityPerformanceCleaner(path=excel_file_path, sheet_name='规模明细')
                # data = cleaner.clean()

                # 创建表格
                mysqldb.create_table(mysql_table_name, data)

                # 创建关系
                # mysqldb.add_foreign_key(table_name)

                # 导入数据
                mysqldb.import_data(mysql_table_name, data)

    # 读取数据库数据并写入Excel表格
    elif selected_sub_process == 3:
        excel_file_path = r"D:\Users\linsuisheng034\Desktop\规模明细.xlsx"
        with MySqlDBManager(database="business") as mysqldb:
            # 加载现有的Excel文件，如果文件不存在则会抛出异常
            with suppress(FileNotFoundError):
                wb = load_workbook(excel_file_path)
            sheet_name = "202412"
            sql = dedent(f"""
                SELECT *
                FROM business.`规模明细` t1
                WHERE t1.`月度` = "2024-12-01";
            """)
            df = mysqldb.get_full_table(sql_statement=sql)

            # 检查并删除已存在的工作表
            if wb and sheet_name in wb.sheetnames:
                del wb[sheet_name]
                wb.save(excel_file_path)

            # 写入数据到Excel
            with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 数据库字段转换
    elif selected_sub_process == 4:
        for db in ["enterprise"]:
            with MySqlDBManager(database=db) as mysqldb:
                # 读取数据
                old_mysql_table_name = "annuity_cname_hub_old"
                data = mysqldb.get_full_table(
                    sql_statement=f"SELECT * FROM `{old_mysql_table_name}`;"
                )
                # 数据清洗
                from common_utils.common_utils import clean_company_name

                data.drop(columns=["id", "_id"], inplace=True, errors="ignore")
                data["年金账户名"] = data["原客户名称"].apply(clean_company_name)
                # 创建表格
                new_mysql_table_name = "annuity_cname_hub"
                # mysqldb.create_table(new_mysql_table_name, data)
                # 导入数据
                mysqldb.import_data(new_mysql_table_name, data)

    # 机构归属统一调整
    elif selected_sub_process == 5:
        with MySqlDBManager(database="mapping") as db_manager:
            table_name = "组织架构"
            # 使用 pandas 从指定表格读取数据
            df = pd.read_sql(f"SELECT * FROM {table_name}", db_manager.engine)

            # 准备参数化的 SQL 语句
            sql_statement = dedent("""
                UPDATE `全量客户` AS t1
                SET t1.`主拓机构代码` = %s, t1.`主拓机构` = %s
                WHERE
                    t1.`主拓机构` IS NULL
                    AND t1.`年金客户类型` = '流失客户'
                    AND t1.`客户名称` LIKE %s;
            """)

            # 批量参数列表
            params_list = []

            for _, row in df.iterrows():
                params = (row["机构代码"], row["机构"], f"%{row['机构']}%")
                params_list.append(params)

            try:
                # 批量执行 SQL 语句
                db_manager.cursor.executemany(sql_statement, params_list)
                db_manager.conn.commit()
                print("数据更新成功。")
            except Exception as e:
                db_manager.conn.rollback()
                print(f"更新数据时出错：{e}")
