import numpy as np
import pandas as pd
import re
import time

from abc import ABC, abstractmethod
from openpyxl import load_workbook

from common_utils.common_utils import parse_to_standard_date, clean_company_name
from data_handler.mappings import (
    BUSINESS_TYPE_CODE_MAPPING,
    COMPANY_ID1_MAPPING,
    COMPANY_ID2_MAPPING,
    COMPANY_ID3_MAPPING,
    COMPANY_ID4_MAPPING,
    COMPANY_ID5_MAPPING,
    COMPANY_BRANCH_MAPPING,
    DEFAULT_PLAN_CODE_MAPPING,
    DEFAULT_PORTFOLIO_CODE_MAPPING,
    PRODUCT_ID_MAPPING,
    PROFIT_METRICS_MAPPING,
)
from logger.logger import logger


class AbstractCleaner(ABC):
    def __init__(self, path):
        self.file_path = path
        self.sheet_name = None
        self.data = []

    def clean(self) -> pd.DataFrame:
        try:
            df = self._clean_method()
            self.data.append(df)
            return df

        except Exception as e:
            class_name = self.__class__.__name__
            logger.error(
                f"Error in class {class_name} cleaning {self.file_path}. Error: {e}"
            )
            return pd.DataFrame()  # 返回空 DataFrame 或抛出异常

    @abstractmethod
    def _clean_method(self) -> pd.DataFrame:
        """对数据进行处理的抽象方法，需要在子类中具体实现。"""
        pass

    def _load_data(self, sheet=None) -> pd.DataFrame:
        try:
            if not (sheet is None) and self.file_path.endswith(".xlsx"):
                return pd.read_excel(
                    self.file_path, sheet_name=sheet, header=0, dtype=str
                )
            elif self.file_path.endswith(".csv"):
                return pd.read_csv(
                    self.file_path, sep="\t", header=0, encoding="UTF-16"
                )
            elif self.file_path.endswith(".xls"):
                return pd.read_csv(
                    self.file_path, sep="\t", header=0, encoding="UTF-16"
                )
            elif self.file_path.endswith(".xlsx"):
                return pd.read_excel(
                    pd.ExcelFile(self.file_path), sheet_name=0, header=0, dtype=str
                )
            else:
                logger.error(f"Unsupported file format or keyword: {self.file_path}")
                return pd.DataFrame()
        except Exception as e:
            self._log_error("Error reading file", e)
            return pd.DataFrame()

    def _get_data_from_file(self):
        if self.file_path.endswith(".csv"):
            try:
                local_data = pd.read_csv(
                    self.file_path, delimiter=",", encoding="utf-8-sig", dtype=str
                )
                local_data.columns = local_data.columns.str.strip()
                return local_data
            except UnicodeDecodeError:
                try:
                    local_data = pd.read_csv(
                        self.file_path, delimiter=",", encoding="cp936", dtype=str
                    )
                    local_data.columns = local_data.columns.str.strip()
                    return local_data
                except UnicodeDecodeError:
                    # Handle or log the error as needed
                    pass
        elif self.file_path.endswith((".xlsx", ".xls")):
            local_data = pd.read_excel(self.file_path, dtype=str)
            return local_data

        return pd.DataFrame()

    def _log_error(self, message, exception=None):
        if exception:
            logger.error(f"{message}. Error: {exception}. File: {self.file_path}")
        else:
            logger.error(f"{message}. File: {self.file_path}")

    def _rename_columns(self, df, column_mapping):
        try:
            return df.rename(columns=column_mapping)
        except Exception as e:
            self._log_error("Error renaming columns", e)
            return pd.DataFrame()

    def _extract_department_code(self, df, column="G代码"):
        try:
            if column in df.columns:
                df["部门代码"] = np.where(
                    df[column] == "总计",
                    "G05SZ99",
                    df[column].str.extract(r"(G[\dA-Z]+)\)")[0],
                )
                df["部门名称"] = np.where(
                    df[column] == "总计", "", df[column].str.extract(r"\)(.+)")[0]
                )
            return df
        except Exception as e:
            self._log_error("Error extracting department code", e)
            return pd.DataFrame()

    def _add_business_type_code(self, df):
        try:
            df["业务类型代码"] = (
                df["业务类型"].str.lower().map(BUSINESS_TYPE_CODE_MAPPING)
            )
            df["业务类型代码"] = df["业务类型代码"].fillna(df["业务类型"])
            return df
        except Exception as e:
            self._log_error("Error adding business type code", e)
            return pd.DataFrame()

    def _remove_columns(self, df, columns=None):
        try:
            if columns is None:
                columns = ["G代码", "RepressRow", "REPRESSROW"]
            df.drop(columns, axis=1, inplace=True, errors="ignore")
            return df
        except Exception as e:
            self._log_error("Error removing columns", e)
            return pd.DataFrame()

    def _update_company_id(
        self,
        df,
        plan_code_col="计划代码",
        customer_name_col="客户名称",
        company_id_col="company_id",
    ):
        """
        根据指定的映射关系更新 DataFrame 中的 company_id 列。

        参数：
        - df: pandas DataFrame，待处理的数据框。
        - plan_code_col: str，计划代码的列名，默认为 '计划代码'。
        - customer_name_col: str，客户名称的列名，默认为 '客户名称'。
        - company_id_col: str，company_id 的列名，默认为 'company_id'。

        返回：
        - pandas DataFrame，更新后的数据框。
        """

        try:
            # 初始映射
            df[company_id_col] = df[plan_code_col].map(COMPANY_ID1_MAPPING)

            # 处理特殊情况，默认值为 '600866980'
            mask = (df[company_id_col].isna() | (df[company_id_col] == "")) & (
                df[customer_name_col].isna() | (df[customer_name_col] == "")
            )
            company_id_from_plan = (
                df[plan_code_col].map(COMPANY_ID3_MAPPING).fillna("600866980")
            )
            df.loc[mask, company_id_col] = company_id_from_plan[mask]

            # 根据客户名称补充 company_id
            mask = df[company_id_col].isna() | (df[company_id_col] == "")
            company_id_from_customer = df[customer_name_col].map(COMPANY_ID4_MAPPING)
            df.loc[mask, company_id_col] = company_id_from_customer[mask]

            return df
        except Exception as e:
            self._log_error("Error removing columns", e)
            return pd.DataFrame()


# 业务数据-系统：规模明细
class AnnuityPerformanceCleaner(AbstractCleaner):
    def __init__(self, path, sheet_name=None):
        super().__init__(path)
        self.sheet_name = "规模明细" if sheet_name is None else sheet_name

    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data(self.sheet_name)
        df.rename(
            columns={
                "机构": "机构名称",
                "计划号": "计划代码",
                "流失（含待遇支付）": "流失(含待遇支付)",
            },
            inplace=True,
        )
        # 利用机构名称修正机构代码
        df["机构代码"] = df["机构名称"].map(COMPANY_BRANCH_MAPPING)
        # 修正日期
        df["月度"] = df["月度"].apply(parse_to_standard_date)
        # 修正计划代码
        df["计划代码"] = df["计划代码"].replace({"1P0290": "P0290", "1P0807": "P0807"})
        df["计划代码"] = df["计划代码"].mask(
            (df["计划代码"].isna() | (df["计划代码"] == ""))
            & (df["计划类型"] == "集合计划"),
            "AN001",
        )
        df["计划代码"] = df["计划代码"].mask(
            (df["计划代码"].isna() | (df["计划代码"] == ""))
            & (df["计划类型"] == "单一计划"),
            "AN002",
        )
        # 替换null或者空为总部代码G00
        df["机构代码"] = df["机构代码"].replace("null", "G00")
        df["机构代码"] = df["机构代码"].fillna("G00")
        # 安全操作: 修正组合编号
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        else:
            df["组合代码"] = df["组合代码"].str.replace("^F", "", regex=True)
        # 默认组合代码设定
        df["组合代码"] = df["组合代码"].mask(
            (df["组合代码"].isna() | (df["组合代码"] == "")),
            df.apply(
                lambda x: "QTAN003"
                if x["业务类型"] in ["职年受托", "职年投资"]
                else DEFAULT_PORTFOLIO_CODE_MAPPING.get(x["计划类型"]),
                axis=1,
            ),
        )
        # 匹配产品线代码
        df["产品线代码"] = df["业务类型"].map(BUSINESS_TYPE_CODE_MAPPING)
        # 修正客户名称
        df["年金账户名"] = df["客户名称"]
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )

        # 步骤一：根据 '年金计划号' 补充 'company_id'
        df["company_id"] = df["计划代码"].map(COMPANY_ID1_MAPPING)

        # 清洗 '集团企业客户号' 列
        df["集团企业客户号"] = df["集团企业客户号"].str.lstrip("C")

        # 步骤二：根据 '集团企业客户号' 补充 'company_id'
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_group = df["集团企业客户号"].map(COMPANY_ID2_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_group[mask]

        # 步骤三：处理特殊情况，默认值为 '600866980'
        mask = (df["company_id"].isna() | (df["company_id"] == "")) & (
            df["客户名称"].isna() | (df["客户名称"] == "")
        )
        company_id_from_plan = (
            df["计划代码"].map(COMPANY_ID3_MAPPING).fillna("600866980")
        )
        df.loc[mask, "company_id"] = company_id_from_plan[mask]

        # 步骤四：根据 '客户名称' 补充 'company_id'
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["客户名称"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]

        # 步骤五：根据 '年金账户名' 补充 'company_id'
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_account = df["年金账户名"].map(COMPANY_ID5_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_account[mask]

        # 删除无效字段
        df.drop(
            columns=[
                "备注",
                "子企业号",
                "子企业名称",
                "集团企业客户号",
                "集团企业客户名称",
            ],
            inplace=True,
            errors="ignore",
        )
        return df


# 业务数据-系统：收入明细
class AnnuityIncomeCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("收入明细")
        df.rename(
            columns={
                "机构": "机构代码",
            },
            inplace=True,
        )
        # 利用机构名称修正机构代码
        df["机构代码"] = df["机构名称"].map(COMPANY_BRANCH_MAPPING)
        # 修正日期
        df["月度"] = df["月度"].apply(parse_to_standard_date)
        # 替换null或者空为总部代码G00
        df["机构代码"] = df["机构代码"].replace("null", "G00")
        df["机构代码"] = df["机构代码"].fillna("G00")
        # 安全操作: 修正组合编号
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        else:
            df["组合代码"] = df["组合代码"].str.replace("^F", "", regex=True)
        # 默认组合代码设定
        df["组合代码"] = df["组合代码"].mask(
            (df["组合代码"].isna() | (df["组合代码"] == "")),
            df.apply(
                lambda x: "QTAN003"
                if x["业务类型"] in ["职年受托", "职年投资"]
                else DEFAULT_PORTFOLIO_CODE_MAPPING.get(x["计划类型"]),
                axis=1,
            ),
        )
        # 匹配产品线代码
        df["产品线代码"] = df["业务类型"].map(BUSINESS_TYPE_CODE_MAPPING)
        # 修正客户名称
        df["年金账户名"] = df["客户名称"]
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )

        df = self._update_company_id(
            df, plan_code_col="计划号", customer_name_col="客户名称"
        )

        # 根据 '年金账户名' 补充 'company_id'
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_account = df["年金账户名"].map(COMPANY_ID5_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_account[mask]

        return df


# 业务数据-系统：团养缴费
class GroupRetirementCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("团养缴费")
        df.rename(columns={"机构": "机构名称"}, inplace=True)
        # 修正日期
        for col in ["缴费期间(起)", "缴费期间(止)", "到账日期", "交易日期", "申请日期"]:
            df[col] = df[col].apply(parse_to_standard_date)
        # 新增机构代码
        if "机构代码" not in df.columns:
            df["机构代码"] = np.nan
        df["机构代码"] = df["机构名称"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["企业名称"] = df["企业名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 根据 '客户名称' 补充 'company_id'
        if "company_id" not in df.columns:
            df["company_id"] = np.nan
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["企业名称"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]
        return df


# 业务数据-系统：企康缴费数据（养健合并）
class HealthCoverageCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("企康缴费")
        # 修正日期
        for col in ["承保时间", "生效时间", "缴费确认时间"]:
            df[col] = df[col].apply(parse_to_standard_date)
        # 新增机构代码
        if "机构代码" not in df.columns:
            df["机构代码"] = np.nan
        # df['机构代码'] = df['机构代码'].str.strip().str[:3]
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        df.rename(
            columns={
                "机构": "机构名称",
                "项目名称": "项目简称",
                "保单号/客户名称": "保单号",
                "um号": "UM号",
            },
            inplace=True,
        )
        # 考核业绩列
        df["考核业绩"] = df["最终业绩"]
        # 安全删除列
        df.drop(
            columns=[
                "所属渠道",
                "战区",
                "中心",
                "渠道",
                "战区",
                "缴费年",
                "缴费月",
                "新/续保",
                "折算后业绩",
                "最终业绩",
            ],
            inplace=True,
            errors="ignore",
        )
        # 修正客户名称
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 根据 '客户名称' 补充 'company_id'
        if "company_id" not in df.columns:
            df["company_id"] = np.nan
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["客户名称"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]
        return df


# 业务数据-系统：企康缴费_养老险承保数据
class YLHealthCoverageCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("养老险")
        # 修正日期
        for col in ["承保时间", "生效时间", "缴费确认时间"]:
            df[col] = df[col].apply(parse_to_standard_date)
        # 新增机构代码
        if "机构代码" not in df.columns:
            df["机构代码"] = np.nan
        # df['机构代码'] = df['机构代码'].str.strip().str[:3]
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        df.rename(columns={"机构": "机构名称"}, inplace=True)
        # 安全删除列
        df.drop(
            columns=[
                "所属渠道",
                "战区",
                "中心",
                "渠道",
                "项目简称",
                "战区",
                "缴费年",
                "缴费月",
                "新/续保",
            ],
            inplace=True,
            errors="ignore",
        )
        # 修正客户名称
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 根据 '客户名称' 补充 'company_id'
        if "company_id" not in df.columns:
            df["company_id"] = np.nan
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["客户名称"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]
        return df


# 业务数据-系统：企康缴费_健康险承保数据
class JKHealthCoverageCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("健康险")
        # 修正日期
        for col in ["资金变动日期", "满期日"]:
            df[col] = df[col].apply(parse_to_standard_date)
        # 新增机构代码
        if "机构代码" not in df.columns:
            df["机构代码"] = np.nan
        df["机构代码"] = df["二级机构"].map(COMPANY_BRANCH_MAPPING)
        # 安全删除列
        del_columns = [
            col
            for col in df.columns
            if "渠道" in col
            or col
            in [
                "拆分比例",
                "币种",
                "战区",
                "中心",
                "缴费年",
                "缴费月",
                "项目起保日期",
                "G代码",
            ]
        ]
        df.drop(columns=del_columns, inplace=True, errors="ignore")
        # 使用正则表达式和捕获组直接替换列名
        df.columns = [
            re.sub(r"(.+)[\(\（](.+)[\)\）]", r"\1_\2", col.replace("\n", ""))
            for col in df.columns
        ]
        df.rename(columns={"缴费任务": "投保单状态"}, inplace=True)
        # 修正客户名称
        df["投保人"] = df["投保人"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 根据 '客户名称' 补充 'company_id'
        if "company_id" not in df.columns:
            df["company_id"] = np.nan
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["投保人"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]
        return df


# 业务数据-台账：提费扩面/increase_fee_and_expand_coverage
class IFECCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("提费扩面")
        df.rename(columns={"机构": "机构名称"}, inplace=True)
        # 修正日期
        df["月度"] = df["月度"].apply(parse_to_standard_date)
        # 新增机构代码
        df["机构代码"] = df["机构名称"].map(COMPANY_BRANCH_MAPPING)
        # 转换'新客固费费率'为数字
        df["新客固费费率"] = df["新客固费费率"].str.replace("%", "").astype(float) / 100
        # 确保数据类型正确
        df["已客提费扩面供款"] = pd.to_numeric(df["已客提费扩面供款"], errors="coerce")
        # 修正客户名称和company_id
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        df = self._update_company_id(
            df, plan_code_col="年金计划号", customer_name_col="客户名称"
        )

        return df


# 业务数据-台账：手工调整/annuity_performance_manual_adjustment
class APMACleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("灌入数据")
        df.rename(columns={"机构": "机构名称"}, inplace=True)
        # 修正日期
        df["月度"] = df["月度"].apply(parse_to_standard_date)
        # 新增机构代码
        df["机构代码"] = df["机构名称"].map(COMPANY_BRANCH_MAPPING)
        # 安全操作: 修正组合编号
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        else:
            df["组合代码"] = df["组合代码"].str.replace("^F", "", regex=True)
        # 默认组合代码设定
        df["组合代码"] = df["组合代码"].mask(
            (df["组合代码"].isna() | (df["组合代码"] == "")),
            df.apply(
                lambda x: "QTAN003"
                if x["业务类型"] in ["职年受托", "职年投资"]
                else DEFAULT_PORTFOLIO_CODE_MAPPING.get(x["计划类型"]),
                axis=1,
            ),
        )
        # 匹配产品线代码
        df["产品线代码"] = df["业务类型"].map(BUSINESS_TYPE_CODE_MAPPING)
        # 修正客户名称
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        df = self._update_company_id(
            df, plan_code_col="计划代码", customer_name_col="客户名称"
        )

        return df


# 业务数据-台账：企年受托中标（中标）
class TrusteeAwardCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("企年受托中标(空白)")
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["客户全称"] = df["客户全称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 删除无效字段
        df.drop(columns=["insert_sql", ""], inplace=True, errors="ignore")
        df.drop(
            columns=[col for col in df.columns if col.startswith("Unnamed")],
            inplace=True,
            errors="ignore",
        )
        return df


# 业务数据-台账：企年受托流失
class TrusteeLossCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("企年受托流失(解约)")
        # 筛选导入数据
        df = df[~df["年金计划号"].isnull()]
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["客户全称"] = df["客户全称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 删除无效字段
        df.drop(
            columns=["insert_sql", "", "Unnamed: 18"], inplace=True, errors="ignore"
        )

        return df


# 业务数据-台账：企年投资中标
class InvesteeAwardCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("企年投资中标(空白)")
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["客户全称"] = df["客户全称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 删除无效字段
        df.drop(columns=["insert_sql", ""], inplace=True, errors="ignore")
        df.drop(
            columns=[col for col in df.columns if col.startswith("Unnamed")],
            inplace=True,
            errors="ignore",
        )
        return df

        return df


# 业务数据-台账：企年投资流失
class InvesteeLossCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("企年投资流失(解约)")
        # 筛选导入数据
        df = df[~df["年金计划号"].isnull()]
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["客户全称"] = df["客户全称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 删除无效字段
        df.drop(
            columns=["insert_sql", "", "Unnamed: 18"], inplace=True, errors="ignore"
        )

        return df


# 业务数据-台账：职年投资新增组合/pension_investee_new_investment_portfolio
class PInvesteeNIPCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("职年投资新增组合")
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        df["到账月份"] = df["到账月份"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)

        return df


# 业务数据-台账：组合业绩
class InvestmentPortfolioCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("组合业绩")
        # 表头清洗
        df.columns = df.columns.str.replace("\n", "", regex=True)  # 替换换行符
        df.columns = df.columns.str.replace("\t", "", regex=True)  # 替换制表符
        # 修正组合编号
        df["组合编号"] = df["组合编号"].str.replace("^F", "", regex=True)
        # 修正日期
        df["数据更新日期"] = df["数据更新日期"].apply(parse_to_standard_date)

        return df


# 业务数据-台账：团养中标
class GRAwardCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("团养中标")
        # 修正日期
        df["上报月份"] = df["上报月份"].apply(parse_to_standard_date)
        df["中标/签约时间"] = df["中标/签约时间"].apply(parse_to_standard_date)
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正客户名称
        df["客户名称"] = df["客户名称"].apply(
            lambda x: clean_company_name(x) if isinstance(x, str) else x
        )
        # 根据 '客户名称' 补充 'company_id'
        if "company_id" not in df.columns:
            df["company_id"] = np.nan
        mask = df["company_id"].isna() | (df["company_id"] == "")
        company_id_from_customer = df["客户名称"].map(COMPANY_ID4_MAPPING)
        df.loc[mask, "company_id"] = company_id_from_customer[mask]
        return df


# 业务数据-台账：续签客户清单
class RenewalPendingCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("续签客户清单")
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)

        return df


# 收入数据：风准金余额表
class RiskProvisionBalanceCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("风准金")
        # 表头清洗
        df.columns = df.columns.str.strip()
        df.columns = df.columns.str.replace("\n", "", regex=True)  # 替换换行符
        df.columns = df.columns.str.replace("\t", "", regex=True)  # 替换制表符
        # 机构代码
        df.rename(columns={"机构": "属地"}, inplace=True)
        df["机构代码"] = df["属地"].map(COMPANY_BRANCH_MAPPING)
        # 修正日期
        df["结转日期"] = df["结转日期"].apply(parse_to_standard_date)
        # 安全操作: 修正组合编号
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        else:
            df["组合代码"] = df["组合代码"].str.replace("^F", "", regex=True)
        # 补充计划代码
        df["计划代码"] = df["组合代码"].map(DEFAULT_PLAN_CODE_MAPPING)
        return df


# 收入数据：历史浮费
class HistoryFloatingFeesCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("历史浮费")
        # 表头清洗
        df.columns = df.columns.str.replace("（元）", "")  # 假设没有正则表达式特殊字符
        df.columns = df.columns.str.strip()  # 去除列名中的前后空格
        df.columns = df.columns.str.replace("\n", "", regex=True)  # 替换换行符
        df.columns = df.columns.str.replace("\t", "", regex=True)  # 替换制表符
        # 筛选出非空且非空字符串的行
        df = df[df["计提月度"].notna() & (df["计提月度"] != "")]
        # 机构代码
        df["机构代码"] = df["机构"].map(COMPANY_BRANCH_MAPPING)
        # 修正日期
        df["计提月度"] = df["计提月度"].apply(parse_to_standard_date)
        # 安全操作: 修正组合编号
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        else:
            df["组合代码"] = df["组合代码"].str.replace("^F", "", regex=True)
        # 向下填充空值
        df[["组合代码", "组合名称", "浮动金额"]] = df[
            ["组合代码", "组合名称", "浮动金额"]
        ].fillna(method="ffill")
        # 补充计划代码
        df["计划代码"] = df["组合代码"].map(DEFAULT_PLAN_CODE_MAPPING)

        # 只保留存在的列
        columns_needed = [
            "区域",
            "机构",
            "受托人",
            "组合代码",
            "组合名称",
            "浮动金额",
            "浮费年度",
            "已计提浮动金额",
            "计提月度",
        ]
        df = df.filter(items=columns_needed)

        return df


# 收入数据：减值计提
class AssetImpairmentCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data("减值计提")
        # 机构代码
        df["机构代码"] = (
            df["分支机构"].str.replace("分公司汇总", "").map(COMPANY_BRANCH_MAPPING)
        )
        # 修正日期
        df["月度"] = df["月度"].apply(parse_to_standard_date)
        # 产品线代码
        df["产品线代码"] = df["业务类型"].map({"投管": "PL201", "受托": "PL202"})

        return df


# 收入数据：考核口径利润达成
class RevenueDetailsCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        dfs = []
        for sheet_name in [
            "公司利润数据"
        ]:  # 原读取表sheet_name'实际_2024', '实际_2023'
            df = self._load_data(sheet_name)
            # 筛选导入数据
            df = df[df["是否导入"].isnull()].drop(columns=["是否导入"])
            df = df.iloc[:, 2:]
            # 数据逆透视
            df = df.melt(
                id_vars=["产品线", "产品明细", "利润指标"],
                var_name=["月份"],
                value_name="指标达成",
            )
            # 日期修正
            year_mapping = {
                "实际_2023": 2023,
                "实际_2024": 2024,
            }
            year = year_mapping.get(sheet_name, pd.Timestamp.now().year)
            df["月份"] = pd.to_datetime(
                df["月份"]
                .str.replace(r"(\d+)月累计", r"\1", regex=True)
                .apply(lambda x: f"{year}-{int(x):02d}-01"),
                format="%Y-%m-%d",
            )
            # 字段映射
            df["产品线代码"] = df["产品线"].map(BUSINESS_TYPE_CODE_MAPPING)
            df["产品ID"] = df["产品明细"].map(PRODUCT_ID_MAPPING)
            df["指标编码"] = df["利润指标"].map(PROFIT_METRICS_MAPPING)
            # 确保 '指标达成' 列为数值类型
            df["指标达成"] = pd.to_numeric(df["指标达成"], errors="coerce")
            # 数值符号修正(Pandas 的 loc 允许我们通过布尔索引选择 DataFrame 的特定部分，并直接对这些部分进行操作)
            df.loc[df["利润指标"] == "收支轧差", "指标达成"] = -df["指标达成"]
            df = df[df["指标达成"].notna() & (df["指标达成"] != "")]

            dfs.append(df)

        return pd.concat(dfs, axis=0)


# 收入数据：考核口径利润预算
class RevenueBudgetCleaner(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        dfs = []
        for sheet_name in ["预算_2024"]:
            df = self._load_data(sheet_name)
            # 筛选导入数据
            df = df[df["是否导入"].isnull()].drop(columns=["是否导入"])
            df = df.iloc[:, 2:]
            # 数据逆透视
            df = df.melt(
                id_vars=["产品线", "产品明细", "利润指标"],
                var_name=["月份"],
                value_name="指标预算",
            )
            # 日期修正
            year = 2023 if sheet_name == "预算_2023" else pd.Timestamp.now().year
            df["月份"] = pd.to_datetime(
                df["月份"]
                .str.replace(r"(\d+)月累计", r"\1", regex=True)
                .apply(lambda x: f"{year}-{int(x):02d}-01"),
                format="%Y-%m-%d",
            )
            # 字段映射
            df["产品线代码"] = df["产品线"].map(BUSINESS_TYPE_CODE_MAPPING)
            df["产品ID"] = df["产品明细"].map(PRODUCT_ID_MAPPING)
            df["指标编码"] = df["利润指标"].map(PROFIT_METRICS_MAPPING)
            # 确保 '指标达成' 列为数值类型
            df["指标预算"] = pd.to_numeric(df["指标预算"], errors="coerce")
            # 数值符号修正(Pandas 的 loc 允许我们通过布尔索引选择 DataFrame 的特定部分，并直接对这些部分进行操作)
            df.loc[df["利润指标"] == "收支轧差", "指标预算"] = -df["指标预算"]
            df = df[df["指标达成"].notna() & (df["指标达成"] != "")]

            dfs.append(df)

        return pd.concat(dfs, axis=0)


# 收入数据：年金费率统计表
class AnnuityRateStatisticsData(AbstractCleaner):
    def _clean_method(self) -> pd.DataFrame:
        df = self._load_data(0)
        df.rename(
            columns={
                "投资/受托": "业务类型",
                "计划/组合名称": "计划组合名称",
                "计划/组合代码": "计划组合代码",
                "2023年底/新签规模": "资产规模",
                "续签/新签合同费率": "合同费率",
                "是否新签/续签": "签约类型",
                "新签/续签": "签约类型",
                "新签/续签时间": "签约时间",
            },
            inplace=True,
        )
        df["计划组合代码"] = df["计划组合代码"].mask(
            (df["计划组合代码"].isna() | (df["计划组合代码"] == ""))
            & (df["计划类型"] == "单一计划"),
            "AN002",
        )
        # 修正日期
        df["签约时间"] = df["签约时间"].apply(parse_to_standard_date)
        # 补充字段
        if "组合代码" not in df.columns:
            df["组合代码"] = np.nan
        if "计划代码" not in df.columns:
            df["计划代码"] = np.nan
        # 补充受托字段
        df["计划代码"] = df["计划代码"].mask(
            df["业务类型"] == "受托", df["计划组合代码"]
        )
        df["组合代码"] = df["组合代码"].mask(
            df["业务类型"] == "受托", df["计划类型"].map(DEFAULT_PORTFOLIO_CODE_MAPPING)
        )
        # 补充投资字段
        df["组合代码"] = df["组合代码"].mask(
            df["业务类型"] == "投资",
            df["计划组合代码"].str.replace("^F", "", regex=True),
        )
        df["计划代码"] = df["计划代码"].mask(
            df["业务类型"] == "投资", df["组合代码"].map(DEFAULT_PLAN_CODE_MAPPING)
        )

        return df


if __name__ == "__main__":
    selected_sub_process = int(input("Please enter an integer: "))

    # 清洗函数测试
    if selected_sub_process == 1:
        start_time = time.time()
        file_path = r"D:\Share\DATABASE\Monthly Update\【for年金分战区经营分析】24年6月年金终稿数据0708采集(手动校准).xlsx"
        cleaner = AnnuityPerformanceCleaner(
            path=file_path, sheet_name="2312企年投资集合计划当月"
        )
        data = cleaner.clean()
        # 写入数据到Excel
        saved_path = r"D:\Users\linsuisheng034\Desktop\temp.xlsx"
        # 指定工作表名称
        sheet_name = "企康缴费_健"
        # 检查并删除已存在的工作表
        wb = load_workbook(saved_path)
        if wb and sheet_name in wb.sheetnames:
            del wb[sheet_name]
            wb.save(saved_path)
        with pd.ExcelWriter(saved_path, engine="openpyxl", mode="a") as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
        # print(data.head(10))
        end_time = time.time()
        print(f"""It has cost {end_time - start_time:.2f} seconds.""")

    # 其他函数测试
    elif selected_sub_process == 2:
        name = "鄂尔多斯市嘉泰矿业有限责任公司（已转出）"
        clean_name = clean_company_name(name)
        print(f"Name: {name}, CleanName: {clean_name}")
